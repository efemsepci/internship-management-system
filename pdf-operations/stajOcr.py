from flask import Flask, request, jsonify
from flask_cors import CORS
import os
from pdf2image import convert_from_path
import easyocr
import pytesseract
from PIL import Image, ImageEnhance, ImageFilter
import numpy as np
import openpyxl

app = Flask(__name__)

CORS(app)

UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': 'Dosya bulunamadı.'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Dosya adı boş.'}), 400

    if file and file.filename.endswith('.pdf'):
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
        file.save(file_path)

        try:
            # PDF işleme fonksiyonunu çağırıyoruz
            # Örneğin: process_pdf() fonksiyonu PDF dosyasını işler ve metinleri döndürür
            information_text = process_pdf(file_path, internship_information["page"], internship_information["left"], internship_information["top"], internship_information["right"],  internship_information["bottom"],internship_information["output_image_path"], internship_information["is_preprocess"], internship_information["ocrlibrary_flag"])

            evaluation_text = process_pdf(file_path, intern_evaluation["page"], intern_evaluation["left"], intern_evaluation["top"], intern_evaluation["right"], intern_evaluation["bottom"], intern_evaluation["output_image_path"], intern_evaluation["is_preprocess"], intern_evaluation["ocrlibrary_flag"])
            evaluation_text = edit_text(evaluation_text)

            remark_text = process_pdf(file_path, intern_remark["page"], intern_remark["left"], intern_remark["top"], intern_remark["right"], intern_remark["bottom"], intern_remark["output_image_path"], intern_remark["is_preprocess"], intern_remark["ocrlibrary_flag"])
            remark_text = combine_text_lines(remark_text)

            history_text = process_pdf(file_path, intern_history["page"], intern_history["left"],  intern_history["top"], intern_history["right"], intern_history["bottom"], intern_history["output_image_path"], intern_history["is_preprocess"], intern_history["ocrlibrary_flag"])
            history_text = combine_text_lines(history_text)

            data = {
                'Internship Information': information_text,
                'Intern Evaluation': evaluation_text,
                'Remark': remark_text,
                'History': history_text
            }
            
            save_to_excel('internship_evaluation.xlsx', data)
            
            return jsonify({'message': 'Dosya başarıyla işlendi ve kaydedildi.'}), 200
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    return jsonify({'error': 'Sadece PDF dosyaları kabul edilir.'}), 400

#convert pdf to jpg
def pdf_to_images(pdf_path):
    try:
        images = convert_from_path(pdf_path)
        print(f"Converted {len(images)} pages.")
        return images
    except Exception as e:
        print(f"Error converting PDF to images: {str(e)}")
        return []


#crop image
def crop_image(image, left, top, right, bottom):
    try:
        cropped_img = image.crop((left, top, right, bottom))
        return cropped_img
    except Exception as e:
        print(f"Error during image cropping: {e}")
        raise  



def preprocess_image(image):
    try:
        # Increase contrast
        enhancer = ImageEnhance.Contrast(image)
        image = enhancer.enhance(4)

        # Resize
        image = image.resize((image.width * 4, image.height * 2))

        # Noise reduction
        image = image.filter(ImageFilter.MedianFilter())

        # Convert image to grayscale
        image = image.convert('L')

        return image
    except Exception as e:
        print(f"Error during image preprocessing: {e}")
        raise 


#we have 2 functions for ocr operations because easyocr is heavy library.
#but pytesseract can't extract text perfectly in some areas so we are using
#easyocr where pytesseract can't work correctly.

#ocr operation with using easyocr
def extract_text_from_image_easyocr(image):
    try:
        reader = easyocr.Reader(['tr'], gpu=True)
        image_np = np.array(image)
        text_lines = reader.readtext(image_np, detail=0)
        return text_lines
    except Exception as e:
        print(f"Error with EasyOCR: {str(e)}")
        return []


#ocr operation with using pytesseract
def extract_text_from_image_pytesseract(image):
    try:
        # PIL Image'ı numpy array'e dönüştürme
        image_np = np.array(image)

        custom_config = r'--oem 3 --psm 11'

        # Tesseract OCR kullanarak metin çıkarma
        text = pytesseract.image_to_string(image_np, lang='tur', config=custom_config)

        # Çıkarılan metni satırlara bölme
        text_lines = text.split('\n')

        # Boş satırları temizleme
        text_lines = [line.strip() for line in text_lines if line.strip()]

        return text_lines

    except Exception as e:
        print(f"Error during OCR processing: {e}")
        return []  # Boş bir liste döndürerek fonksiyonun güvenli çalışmasını sağlıyoruz


# all function in processs_pdf function
def process_pdf(pdf_path, page_num, left, top, right, bottom, output_image_path=None, process = False, ocr_flag = None):
    # PDF to image
    images = pdf_to_images(pdf_path)

    # take specific pdf page
    page_image = images[page_num - 1]  # starting 0

    width, height = page_image.size

    # crop specific area
    cropped_image = crop_image(page_image, left, top, right, bottom)

    # increase image quality
    if process:
        cropped_image = preprocess_image(cropped_image)

    if output_image_path:
        cropped_image.save(output_image_path)

    # take text from image
    if ocr_flag == 0:
        text_lines = extract_text_from_image_easyocr(cropped_image)
    elif ocr_flag == 1:
        text_lines = extract_text_from_image_pytesseract(cropped_image)
    
    return text_lines

def combine_text_lines(text_lines):
    combined_text = ' '.join(text_lines)
    return combined_text

# #to combine seperate text
# def combine_specific_line(text_lines):
#     text_lines[7] = text_lines[7] + text_lines[8]
#     text_lines.pop(8)
#     return text_lines

#in evaluation area it can detect Id instead of id
def edit_text(text):
    edited_lines = []
    if isinstance(text, list):
        for line in text:
            if 'Id' in line:
                line = line.replace('Id', 'id')
            edited_lines.append(line)
        return edited_lines
    
 #save extracted text to excel file
def save_to_excel(file_name, data):
    try:
        workbook = openpyxl.load_workbook(file_name)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    try:
        sheet = workbook.active
        start_row = sheet.max_row + 1 if sheet.max_row >= 1 else 1
        current_col = 1
        for text_lines in data.values():
            if isinstance(text_lines, list):
                for line in text_lines:
                    sheet.cell(row=start_row, column=current_col, value=line)
                    current_col += 1
            else:
                sheet.cell(row=start_row, column=current_col, value=text_lines)
                current_col += 1
        workbook.save(file_name)
        print(f"Excel file saved: {file_name}")
    except Exception as e:
        print(f"Error saving to Excel: {str(e)}")
        return jsonify({'error': f"Error saving to Excel: {str(e)}"}), 500



#informations about cropped images. output_image_path is just for checking.
#flag value 0 for easyocr 1 for pytesseract
internship_information = {
    "page" : 1,
    "left" : 862,
    "top" : 800,
    "right": 1560,
    "bottom": 1900,
    "output_image_path": None,
    "is_preprocess" : False,
    "ocrlibrary_flag" : 1
}
intern_evaluation = {
    "page" : 2,
    "left" : 1280,
    "top" : 400,
    "right" : 1420,
    "bottom" : 1920,
    "output_image_path" : None,
    "is_preprocess" : True,
    "ocrlibrary_flag" : 0
}
intern_remark = {
    "page" : 2,
    "left" : 200,
    "top" : 1950,
    "right" : 1554,
    "bottom" : 2100,
    "output_image_path" : None,
    "is_preprocess" : False,
    "ocrlibrary_flag" : 0
}
intern_history = {
    "page" : 2,
    "left" : 1300,
    "top" : 2100,
    "right" : 1600,
    "bottom" : 2170,
    "output_image_path" : None,
    "is_preprocess" : False,
    "ocrlibrary_flag" : 1
}

if __name__ == '__main__':
    app.run(debug=True, port=5000)
